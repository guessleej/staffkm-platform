"""文檔 / 知識庫匯出（Round 10-3）。

兩種輸出：
- to_excel_bytes(docs)        匯出文檔 metadata 表（Excel）
- to_zip_bytes(docs, get_para) 匯出文檔內容 ZIP（每篇一個 markdown）

格式
- Excel 欄位：id / name / file_type / file_size / status / paragraphs / chars / tags / hit_strategy / enabled / created_at
- ZIP 結構：
    documents/
      ├ {doc-name}.md      ← 段落 content 串成 markdown
      └ {doc-name}.meta.json
    questions.json         ← 文檔級 questions 彙整

純 stdlib + openpyxl（已在 runtime image）。
"""
from __future__ import annotations

import io
import json
import zipfile
from typing import Any, Callable


def kb_full_export_zip(
    kb_meta: dict[str, Any],
    docs: list[dict[str, Any]],
    paragraphs_by_doc: dict[str, list[dict[str, Any]]],
    *,
    embeddings_by_paragraph: dict[str, list[float]] | None = None,
) -> bytes:
    """v2.8 H1：整 KB 帶 metadata 匯出（與 to_zip_bytes 用於 doc 內容的場景不同）。

    ZIP 結構：
      kb.json                     ← KnowledgeBase 全 meta（name / embedding_model / chunk_* / source_* / settings）
      documents.json              ← 全部 Document 的 metadata + tags + hit_strategy + is_enabled
      paragraphs/{doc_id}.json    ← 每文件的 paragraphs 全 metadata（含 qa_pairs / order_index）
      embeddings/{paragraph_id}.b64  ← 可選；?include_embeddings=1 時帶
    """
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("kb.json", json.dumps(kb_meta, ensure_ascii=False, indent=2, default=str))
        zf.writestr("documents.json",
                    json.dumps(docs, ensure_ascii=False, indent=2, default=str))
        for doc_id, paras in paragraphs_by_doc.items():
            zf.writestr(
                f"paragraphs/{doc_id}.json",
                json.dumps(paras, ensure_ascii=False, indent=2, default=str),
            )
        if embeddings_by_paragraph:
            import base64
            import struct
            for pid, vec in embeddings_by_paragraph.items():
                # float32 packed → base64，比 JSON list 小 ~6x
                buf = b"".join(struct.pack("<f", float(x)) for x in vec)
                zf.writestr(
                    f"embeddings/{pid}.b64",
                    base64.b64encode(buf).decode("ascii"),
                )
    return out.getvalue()


def _slugify(name: str, idx: int) -> str:
    safe = "".join(c if c.isalnum() or c in "-_." else "_" for c in (name or "doc"))
    return f"{idx:03d}_{safe[:60]}"


def to_excel_bytes(docs: list[dict]) -> bytes:
    """文檔 metadata → Excel；docs 是 list of dict（與 list_documents 回傳格式一致）。"""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"

    headers = [
        "id", "name", "file_type", "file_size", "status",
        "paragraph_count", "char_count", "tags", "hit_strategy",
        "is_enabled", "created_at",
    ]
    ws.append(headers)
    for d in docs:
        ws.append([
            str(d.get("id", "")),
            d.get("name", ""),
            d.get("file_type", ""),
            d.get("file_size", 0),
            d.get("status", ""),
            d.get("paragraph_count", 0),
            d.get("char_count", 0),
            ",".join(d.get("tags", []) or []),
            d.get("hit_strategy", "rag"),
            "✓" if d.get("is_enabled") else "",
            d.get("created_at", ""),
        ])
    # 自動寬度（簡化）
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_zip_bytes(
    docs: list[dict],
    get_paragraphs: Callable[[str], list[dict[str, Any]]],
    *,
    questions_aggregate: list[str] | None = None,
) -> bytes:
    """文檔內容 → ZIP；get_paragraphs(doc_id) 回 [{order_index, content, title?}]。"""
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        for idx, d in enumerate(docs, start=1):
            slug = _slugify(d.get("name", ""), idx)
            paras = get_paragraphs(str(d["id"]))
            # markdown
            md_lines: list[str] = [f"# {d.get('name', slug)}", ""]
            for p in paras:
                t = (p.get("title") or "").strip()
                if t:
                    md_lines.append(f"## {t}")
                    md_lines.append("")
                md_lines.append(p.get("content", "").strip())
                md_lines.append("")
            zf.writestr(f"documents/{slug}.md", "\n".join(md_lines))
            # meta
            meta = {
                "id":              str(d.get("id", "")),
                "name":             d.get("name", ""),
                "file_type":        d.get("file_type", ""),
                "tags":             d.get("tags", []) or [],
                "hit_strategy":     d.get("hit_strategy", "rag"),
                "is_enabled":       bool(d.get("is_enabled", True)),
                "paragraph_count":  d.get("paragraph_count", 0),
                "char_count":       d.get("char_count", 0),
                "questions":        d.get("questions", []) or [],
            }
            zf.writestr(
                f"documents/{slug}.meta.json",
                json.dumps(meta, ensure_ascii=False, indent=2),
            )
        if questions_aggregate:
            zf.writestr(
                "questions.json",
                json.dumps({"questions": questions_aggregate},
                            ensure_ascii=False, indent=2),
            )
    return out.getvalue()
